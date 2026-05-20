"""Thread-scoped file manager orchestration.

Five operations on the per-thread virtual filesystem rooted at the
agent's workspace prefix (``enterprise/{eid}/workspaces/{tid}/``) — the
*same* keys the agent's ``S3Backend`` reads/writes and the Drive mirror
syncs. So a paperclip upload here is immediately visible to the agent
and pushed to the user's Drive folder on the next ``sync_out``.

Operations:

* ``list_files`` — list under a relative prefix.
* ``read_file`` — read one file's contents (UTF-8 text).
* ``write_file`` — write/replace one file from a JSON body.
* ``upload_files`` — multipart batch upload into a folder, preserving filenames.
* ``delete_file`` — delete one file.
* ``scaffold_folders`` — drop ``.keep`` placeholder files so newly
  created threads have a discoverable folder layout.

All paths returned to the caller are *thread-relative* — the
``enterprise/{eid}/workspaces/{tid}/`` storage prefix is added on the
way down to the storage adapter and stripped on the way back up.
Permission checks go through ``FilePolicy`` (actor=``"user"`` for
everything reachable from the REST surface; the agent's tools will
call the same handlers with ``actor="agent"``).
"""

import asyncio
import io
import logging
from pathlib import Path as _Path

from fastapi import HTTPException, UploadFile

from backend.infra.registry import get_storage
from backend.schemas.files import FileContent, FileObject, WriteResult
from backend.services.agent.workspace import workspace_prefix
from backend.services.file_policy import FilePolicy, READ, WRITE

logger = logging.getLogger(__name__)

# Names hidden from list responses — internal placeholders for empty
# folders. Drive mirror's ``sync_out`` also skips these by name.
_HIDDEN_NAMES = frozenset({".keep"})

# Folders scaffolded on thread creation. ``.keep`` markers go in each so
# the layout is discoverable via ``list_files`` immediately, before any
# real content lands.
_SCAFFOLD_FOLDERS = ("input/userUpload", "output")


def _validate_id(thread_id: str, enterprise_id: str) -> None:
    """Reject malformed ids before they reach the storage layer."""
    for label, value in (("thread_id", thread_id), ("enterprise_id", enterprise_id)):
        if not value or "/" in value or ".." in value:
            raise HTTPException(status_code=400, detail=f"Invalid {label}: {value!r}")


def _to_storage_key(enterprise_id: str, thread_id: str, rel_path: str) -> str:
    """Turn a thread-relative path into a full storage key."""
    _validate_id(thread_id, enterprise_id)
    return f"{workspace_prefix(enterprise_id, thread_id)}/{rel_path.lstrip('/')}"


def _from_storage_key(enterprise_id: str, thread_id: str, storage_key: str) -> str:
    """Strip the workspace prefix from a storage key for client display."""
    root = workspace_prefix(enterprise_id, thread_id) + "/"
    return storage_key[len(root) :] if storage_key.startswith(root) else storage_key


async def list_files(
    enterprise_id: str,
    thread_id: str,
    prefix: str = "",
    actor: str = "user",
) -> list[FileObject]:
    """List every file beneath ``prefix`` (thread-relative).

    The empty prefix lists every file in the thread, filtered down to
    items the actor has read access to. ``.keep`` placeholder files
    are hidden from the response.

    Args:
        enterprise_id: Caller's tenant id.
        thread_id: The thread's id (URL segment).
        prefix: Thread-relative folder prefix. Empty means thread root.
        actor: ``"user"`` or ``"agent"``.

    Returns:
        Files matching ``prefix`` that the actor has read access to.
    """
    _validate_id(thread_id, enterprise_id)
    storage = get_storage()
    storage_prefix = workspace_prefix(enterprise_id, thread_id)
    if prefix:
        # Permission check on the folder being listed; only enforced when
        # the caller targets a specific folder. The empty-prefix case
        # falls through to per-item filtering below.
        FilePolicy.check(prefix, actor, READ)
        storage_prefix = f"{storage_prefix}/{prefix.lstrip('/').rstrip('/')}/"

    raw = storage.list_objects(storage_prefix)
    out: list[FileObject] = []
    for obj in raw:
        rel = _from_storage_key(enterprise_id, thread_id, obj["key"])
        if rel.rsplit("/", 1)[-1] in _HIDDEN_NAMES:
            continue
        # Filter out items the actor can't read — covers the empty-prefix
        # case where we listed the entire thread.
        if not FilePolicy.can(rel, actor, READ):
            continue
        out.append(
            FileObject(
                key=rel,
                size=obj["size"],
                modified_at=obj["modified_at"],
            )
        )
    return out


async def read_file(
    enterprise_id: str, thread_id: str, path: str, actor: str = "user"
) -> FileContent:
    """Read one file's contents as UTF-8 text.

    Args:
        enterprise_id: Caller's tenant id.
        thread_id: The thread's id.
        path: Thread-relative path of the file to read.
        actor: ``"user"`` or ``"agent"``.

    Returns:
        ``FileContent`` with the relative key, body, and size.

    Raises:
        HTTPException: 400 on invalid path; 403 on policy denial; 404
            if the object doesn't exist.
    """
    FilePolicy.check(path, actor, READ)
    storage = get_storage()
    key = _to_storage_key(enterprise_id, thread_id, path)
    if not storage.exists(key):
        raise HTTPException(status_code=404, detail=f"File not found: {path}")
    try:
        content = storage.read_text(key)
        return FileContent(key=path, content=content, size=len(content.encode("utf-8")))
    except UnicodeDecodeError:
        import base64 as _b64
        raw = storage.read(key)
        return FileContent(
            key=path,
            content=_b64.b64encode(raw).decode("ascii"),
            size=len(raw),
            is_binary=True,
        )


async def write_file(
    enterprise_id: str,
    thread_id: str,
    path: str,
    content: str,
    actor: str = "user",
) -> WriteResult:
    """Write or replace one file from a text body.

    Args:
        enterprise_id: Caller's tenant id.
        thread_id: The thread's id.
        path: Thread-relative destination path.
        content: UTF-8 text to write. Existing object is overwritten.
        actor: ``"user"`` or ``"agent"``.

    Returns:
        ``WriteResult`` with the relative key and final size.

    Raises:
        HTTPException: 400 on invalid path; 403 on policy denial.
    """
    FilePolicy.check(path, actor, WRITE)
    storage = get_storage()
    key = _to_storage_key(enterprise_id, thread_id, path)
    storage.write_text(key, content)
    return WriteResult(key=path, size=len(content.encode("utf-8")))


async def upload_files(
    enterprise_id: str,
    thread_id: str,
    folder: str,
    files: list[UploadFile],
    actor: str = "user",
) -> list[WriteResult]:
    """Persist multipart uploads into ``folder`` (thread-relative).

    Each part's filename is preserved; collisions overwrite. Use this
    instead of PUT for binary content (PDF, DOCX, images).

    After saving, scaffolds ``workspace/parsed/`` (if absent) and calls
    ``parsed_upload`` so downstream parsing can process the new files.

    Args:
        enterprise_id: Caller's tenant id.
        thread_id: The thread's id.
        folder: Thread-relative destination folder.
        files: Multipart parts received by the router.
        actor: ``"user"`` or ``"agent"``.

    Returns:
        One ``WriteResult`` per saved file, in input order.

    Raises:
        HTTPException: 400 on invalid folder/filename; 403 on policy denial.
    """
    folder = folder.strip("/")
    out: list[WriteResult] = []
    for f in files:
        filename = (f.filename or "").strip("/")
        if not filename or "/" in filename or ".." in filename:
            raise HTTPException(
                status_code=400, detail=f"Invalid filename: {filename!r}"
            )
        rel = f"{folder}/{filename}" if folder else filename
        FilePolicy.check(rel, actor, WRITE)
        storage = get_storage()
        # ``UploadFile.read`` returns the full body — fine for chat-sized
        # attachments. Switch to streaming for large files later.
        body = await f.read()
        storage.write(_to_storage_key(enterprise_id, thread_id, rel), body)
        out.append(WriteResult(key=rel, size=len(body)))

    # --- Scaffold workspace/parsed/ and hand off to parser ------------------
    storage = get_storage()
    parsed_keep = _to_storage_key(enterprise_id, thread_id, "workspace/parsed/.keep")
    if not storage.exists(parsed_keep):
        storage.write(parsed_keep, b"")

    await parsed_upload(enterprise_id, thread_id, folder, files, actor)

    return out


async def parsed_upload(
    enterprise_id: str,
    thread_id: str,
    folder: str,
    files: list[UploadFile],
    actor: str = "user",
) -> None:
    """Parse uploaded files into a bronze-style bundle under ``workspace/parsed/{stem}/``.

    Mirrors the layout that ``services.ingestion.store.store_binary`` produces
    for web-fetched files, so the agent sees the same structure regardless of
    whether content arrived via URL fetch or direct upload:

    .. code-block:: text

        workspace/parsed/{stem}/
            original.{ext}          raw bytes (copy of the uploaded file)
            content.md              extracted/converted text
            images/                 (PDF only) embedded images
                page_1_img_0.png
                ...
            meta.json               filename, page count, image list, status

    Reads each file's bytes back from storage (``UploadFile`` bodies already
    consumed by ``upload_files``). Failures for individual files are logged and
    skipped; the rest continue.

    Args:
        enterprise_id: Caller's tenant id.
        thread_id: The thread's id.
        folder: Thread-relative source folder the files were uploaded into.
        files: The same ``UploadFile`` parts passed to ``upload_files``
            (bodies already consumed; bytes are re-read from storage here).
        actor: ``"user"`` or ``"agent"``.
    """
    folder = folder.strip("/")
    storage = get_storage()

    for f in files:
        filename = (f.filename or "").strip("/")
        if not filename:
            continue

        src_key = _to_storage_key(enterprise_id, thread_id, f"{folder}/{filename}")
        if not storage.exists(src_key):
            logger.warning("parsed_upload: source missing %s", src_key)
            continue

        body = storage.read(src_key)
        suffix = _Path(filename).suffix.lower()
        stem = _Path(filename).stem
        bundle_dir = f"workspace/parsed/{stem}"

        try:
            bundle = await asyncio.to_thread(_extract_to_bundle, body, filename, suffix)
        except Exception:
            logger.exception("parsed_upload: extraction failed for %s", filename)
            continue

        # --- Write original file ---------------------------------------------
        storage.write(
            _to_storage_key(enterprise_id, thread_id, f"{bundle_dir}/original{suffix}"),
            body,
        )

        # --- Write content.md ------------------------------------------------
        if bundle["content"]:
            storage.write_text(
                _to_storage_key(enterprise_id, thread_id, f"{bundle_dir}/content.md"),
                bundle["content"],
            )

        # --- Write images (PDF only) -----------------------------------------
        for img_path, img_bytes in bundle.get("images", {}).items():
            storage.write(
                _to_storage_key(enterprise_id, thread_id, f"{bundle_dir}/{img_path}"),
                img_bytes,
            )

        # --- Write meta.json -------------------------------------------------
        storage.write_text(
            _to_storage_key(enterprise_id, thread_id, f"{bundle_dir}/meta.json"),
            bundle["meta"],
        )

        logger.info("parsed_upload: wrote bundle %s", bundle_dir)


def _extract_to_bundle(body: bytes, filename: str, suffix: str) -> dict:
    """Extract content from ``body`` and return a bundle dict. Runs in a thread.

    Args:
        body: Raw file bytes.
        filename: Original filename (used in meta).
        suffix: Lowercase file extension including the dot (e.g. ``".pdf"``).

    Returns:
        Dict with keys:
        - ``content`` (str): Extracted markdown text.
        - ``images`` (dict[str, bytes]): Relative image paths → raw bytes (PDF only).
        - ``meta`` (str): JSON-encoded metadata string.
    """
    import json as _json
    from datetime import datetime, timezone

    content = ""
    images: dict[str, bytes] = {}
    meta: dict = {
        "filename": filename,
        "source_type": suffix.lstrip("."),
        "content_length": len(body),
        "extracted_at": datetime.now(timezone.utc).isoformat(),
        "extraction_status": "failed",
    }

    if suffix == ".pdf":
        from backend.services.ingestion.store import _extract_pdf_sync

        full_text, page_count, raw_images = _extract_pdf_sync(body)
        meta["pages"] = page_count

        if full_text:
            content = full_text
            meta["extraction_status"] = "extracted"
        elif page_count > 0:
            meta["extraction_status"] = "image_only"

        image_meta: list[dict] = []
        for pg, idx, ext, img_bytes, w, h in raw_images:
            rel = f"images/page_{pg}_img_{idx}.{ext}"
            images[rel] = img_bytes
            image_meta.append({"page": pg, "index": idx, "path": rel, "ext": ext,
                                "width": w, "height": h})
        meta["images"] = image_meta

    elif suffix in (".docx", ".doc"):
        import mammoth

        result = mammoth.convert_to_markdown(io.BytesIO(body))
        content = result.value
        meta["extraction_status"] = "extracted" if content.strip() else "empty"

    elif suffix in (".xlsx", ".xls"):
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(body), read_only=True, data_only=True)
        sheet_parts: list[str] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            sheet_parts.append(f"## {sheet_name}\n")
            header = [str(c or "") for c in rows[0]]
            sheet_parts.append("| " + " | ".join(header) + " |")
            sheet_parts.append("| " + " | ".join("---" for _ in header) + " |")
            for row in rows[1:]:
                sheet_parts.append("| " + " | ".join(str(c or "") for c in row) + " |")
            sheet_parts.append("")
        content = "\n".join(sheet_parts)
        meta["sheets"] = wb.sheetnames
        meta["extraction_status"] = "extracted" if content.strip() else "empty"

    else:
        logger.debug("_extract_to_bundle: unsupported extension %r", suffix)
        meta["extraction_status"] = "unsupported"

    return {"content": content, "images": images, "meta": _json.dumps(meta, indent=2)}


async def delete_file(
    enterprise_id: str, thread_id: str, path: str, actor: str = "user"
) -> None:
    """Delete one file.

    Args:
        enterprise_id: Caller's tenant id.
        thread_id: The thread's id.
        path: Thread-relative path of the file to delete.
        actor: ``"user"`` or ``"agent"``.

    Raises:
        HTTPException: 400 on invalid path; 403 on policy denial; 404
            if the object doesn't exist.
    """
    FilePolicy.check(path, actor, WRITE)
    storage = get_storage()
    key = _to_storage_key(enterprise_id, thread_id, path)
    if not storage.exists(key):
        raise HTTPException(status_code=404, detail=f"File not found: {path}")
    # ``LocalStorage`` doesn't yet expose a ``delete`` method — reach
    # through ``abs_path`` for now. When ``BotoS3Storage`` lands, add a
    # proper delete method to the adapter and switch this caller.
    from pathlib import Path

    Path(storage.abs_path(key)).unlink()


async def scaffold_folders(enterprise_id: str, thread_id: str) -> None:
    """Create the standard folder layout for a freshly created thread.

    Drops a ``.keep`` placeholder under each scaffold folder so the
    layout is discoverable via ``list_files`` before any real content
    lands. ``.keep`` files are hidden from list responses and excluded
    from the Drive mirror's ``sync_out``.

    Args:
        enterprise_id: Caller's tenant id.
        thread_id: The thread's id.
    """
    _validate_id(thread_id, enterprise_id)
    storage = get_storage()
    for folder in _SCAFFOLD_FOLDERS:
        key = _to_storage_key(enterprise_id, thread_id, f"{folder}/.keep")
        if not storage.exists(key):
            storage.write(key, b"")
